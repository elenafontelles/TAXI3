"""Parse Uber daily summary XLSX export into normalized daily summary dicts."""
from __future__ import annotations
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook


def detect_uber_file(filepath: str) -> bool:
    """Detect if file is an Uber daily summary XLSX export."""
    if not filepath.endswith(".xlsx"):
        return False
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        # Check for expected header columns in row 1
        headers = [str(sheet.cell(1, c).value or "").strip().lower() for c in range(1, 9)]
        wb.close()
        return "licencia" in headers and "dia" in headers
    except Exception:
        return False


def _parse_decimal(value) -> Decimal:
    """Parse a cell value into Decimal, handling None and string formats."""
    if value is None:
        return Decimal("0.00")
    s = str(value).strip().replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _parse_date(value) -> date | None:
    """Parse a cell value into a date object."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_uber_xlsx(filepath: str) -> list[dict]:
    """Parse Uber daily summary XLSX export.

    Expected columns (row 1 = headers, row 2+ = data):
    Licencia, Dia, Ganancias totales, Taximetro, Reembolso, Ajustes,
    Total T3 fija uber, Total pago uber

    Returns a list of dicts ready for UberDailySummary model creation.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Read headers from row 1 and build column index
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(1, col).value
        if val:
            headers[str(val).strip().lower()] = col

    # Map expected header names to column indices
    col_map = {
        "licencia": headers.get("licencia"),
        "dia": headers.get("dia"),
        "ganancias_totales": headers.get("ganancias totales"),
        "taximetro": headers.get("taximetro") or headers.get("tax\u00edmetro"),
        "reembolso": headers.get("reembolso"),
        "ajustes": headers.get("ajustes"),
        "t3_fixed": headers.get("total t3 fija uber"),
        "total_payment": headers.get("total pago uber"),
    }

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        row_values = {col: row[idx - 1].value if idx and idx <= len(row) else None
                      for col, idx in col_map.items()}

        license_number = str(row_values.get("licencia") or "").strip()
        if not license_number:
            continue

        day = _parse_date(row_values.get("dia"))
        if not day:
            continue

        t3_fixed = _parse_decimal(row_values.get("t3_fixed"))
        total_payment = _parse_decimal(row_values.get("total_payment"))

        records.append({
            "date": day,
            "license_number": license_number,
            "total_earnings": _parse_decimal(row_values.get("ganancias_totales")),
            "taximeter": _parse_decimal(row_values.get("taximetro")),
            "refund": _parse_decimal(row_values.get("reembolso")),
            "adjustments": _parse_decimal(row_values.get("ajustes")),
            "t3_fixed": t3_fixed,
            "total_payment": total_payment,
        })

    wb.close()
    return records
