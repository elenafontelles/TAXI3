"""Parse Solred/Repsol fuel XLSX reports into normalized expense dicts."""
from __future__ import annotations
import re
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook


def detect_repsol_file(filepath: str) -> bool:
    """Detect if file is a Solred/Repsol fuel report.

    Returns True if file is an XLSX with expected header columns.
    """
    if not filepath.lower().endswith(".xlsx"):
        return False
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        first_row = [str(c.value or "").lower() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        wb.close()
        return "fecha" in first_row and "importe" in first_row
    except Exception:
        return False


def _parse_liters(value) -> Decimal:
    """Parse liters value like '39,120 l' or '43.81'."""
    if value is None:
        return Decimal("0")
    s = str(value).strip().lower().replace(" l", "").replace("l", "")
    # Handle comma as decimal separator
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.001"))
    except Exception:
        return Decimal("0")


def _parse_amount(value) -> Decimal:
    """Parse amount value."""
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    s = str(value).strip().replace("\xa0", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def parse_repsol_pdf(filepath: str) -> list[dict]:
    """Parse Solred/Repsol fuel XLSX report.

    Format (row 1 = headers):
    Fecha | Cuenta | Matrícula | Tarjeta o dispositivo | Producto | Cantidad | Estación | Pago Waylet | Importe

    Kept function name for backwards compatibility with upload route.

    Returns a list of dicts ready for FuelExpense model creation.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    # Read headers to determine column positions
    headers = {}
    for i, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            headers[str(cell.value).strip().lower()] = i

    col_fecha = headers.get("fecha", 0)
    col_plate = headers.get("matrícula", 2)
    col_liters = headers.get("cantidad", 5)
    col_amount = headers.get("importe", 8)

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Parse date
        fecha_val = row[col_fecha] if col_fecha < len(row) else None
        if fecha_val is None:
            continue
        if isinstance(fecha_val, datetime):
            expense_date = fecha_val.date()
        elif isinstance(fecha_val, date):
            expense_date = fecha_val
        else:
            continue

        # Parse plate
        plate_val = row[col_plate] if col_plate < len(row) else None
        plate = str(plate_val or "").strip()
        if not plate or plate == "---":
            continue

        # Parse liters and amount
        liters = _parse_liters(row[col_liters] if col_liters < len(row) else None)
        amount = _parse_amount(row[col_amount] if col_amount < len(row) else None)

        if amount <= 0:
            continue

        records.append({
            "date": expense_date,
            "liters": liters,
            "amount": amount,
            "provider": "solred",
            "payment_method": "tarjeta",
            "_plate": plate,
        })

    wb.close()
    return records
