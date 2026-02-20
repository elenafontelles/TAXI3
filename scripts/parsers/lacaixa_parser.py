"""Parse La Caixa bank statement XLS/XLSX into daily TPV totals."""
from __future__ import annotations
import re
from datetime import datetime, date, timedelta
from decimal import Decimal

# Terminal prefix (first 2 digits) to license number mapping
TERMINAL_LICENSE_MAP = {
    "34": "092",
    "35": "1061",
    "36": "361",
}


def _open_workbook(filepath: str):
    """Open XLS or XLSX and return a uniform row iterator.

    Returns (rows, close_fn) where rows is a list of lists (row 0-indexed),
    and close_fn should be called when done.
    """
    if filepath.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        return rows, wb.close

    # XLS format (xlrd)
    import xlrd
    wb = xlrd.open_workbook(filepath)
    sheet = wb.sheet_by_index(0)
    rows = []
    for i in range(sheet.nrows):
        rows.append([sheet.cell_value(i, j) for j in range(sheet.ncols)])
    return rows, lambda: None


def _excel_serial_to_date(value) -> date | None:
    """Convert Excel serial number to date."""
    if isinstance(value, (int, float)) and value > 30000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(value))).date()
    return None


def detect_lacaixa_file(filepath: str) -> bool:
    """Detect if file is a La Caixa bank statement extract."""
    if not filepath.lower().endswith((".xls", ".xlsx")):
        return False
    try:
        rows, close_fn = _open_workbook(filepath)
        first_cell = rows[0][0] if rows else None
        close_fn()
        return first_cell is not None and "Moviments del compte" in str(first_cell)
    except Exception:
        return False


def _parse_date(value) -> date | None:
    """Parse a cell value into a date object."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    # Excel serial number (XLS format stores dates as floats)
    serial = _excel_serial_to_date(value)
    if serial:
        return serial
    if value is None:
        return None
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value) -> Decimal | None:
    """Parse an amount value, handling comma decimals and EUR suffix."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    s = str(value).strip().replace(" EUR", "").replace("\xa0", "")
    # Handle comma as decimal separator (European format)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except Exception:
        return None


def parse_lacaixa_xlsx(filepath: str) -> list[dict]:
    """Parse La Caixa bank statement XLS/XLSX extract.

    Format:
    - Row 0: Title ("Moviments del compte...")
    - Row 2: Headers (Data, Data valor, Moviment, Mes dades, Import, Saldo)
    - Row 3+: Data rows

    Filters rows where "Moviment" starts with ON or C. followed by known terminal.
    Maps terminal prefixes to license numbers.

    Returns a list of dicts with date, license_number, amount.
    """
    rows, close_fn = _open_workbook(filepath)

    # Read headers from row 2 (0-indexed) to determine column positions
    headers = {}
    if len(rows) > 2:
        for col, val in enumerate(rows[2]):
            if val:
                headers[str(val).strip().lower()] = col

    # Determine column indices (fallback to known positions)
    col_date = headers.get("data", 0)
    col_moviment = headers.get("moviment", 2)
    col_import = headers.get("import", 4)

    records = []
    for row in rows[3:]:  # Data starts at row 3 (0-indexed)
        moviment_val = row[col_moviment] if col_moviment < len(row) else None
        if not moviment_val:
            continue

        moviment_str = str(moviment_val).strip()

        # Match ON or C. entries (TPV card payments and corrections)
        # Format: "ON 363460767 DDMM" or "C. 363460767 DDMM"
        # The first 2 digits of the terminal number determine the license
        license_number = None
        if moviment_str.startswith("ON ") or moviment_str.startswith("C. "):
            parts = moviment_str.split()
            if len(parts) >= 2:
                terminal = parts[1]
                prefix_2 = terminal[:2]
                license_number = TERMINAL_LICENSE_MAP.get(prefix_2)

        if not license_number:
            continue

        # Parse date: extract DDMM from description (last 4 digits),
        # year from "Data" column. E.g. "ON 340229632 2712" -> day=27, month=12
        date_val = row[col_date] if col_date < len(row) else None
        base_date = _parse_date(date_val)
        if not base_date:
            continue

        # Extract DDMM from the moviment description
        day = None
        ddmm_match = re.search(r'\b(\d{4})$', moviment_str)
        if ddmm_match:
            ddmm = ddmm_match.group(1)
            dd, mm = int(ddmm[:2]), int(ddmm[2:])
            try:
                day = date(base_date.year, mm, dd)
            except ValueError:
                day = base_date
        else:
            day = base_date

        if not day:
            continue

        # Parse amount — only positive amounts are TPV income
        import_val = row[col_import] if col_import < len(row) else None
        amount = _parse_amount(import_val)
        if amount is None or amount <= 0:
            continue

        records.append({
            "date": day,
            "license_number": license_number,
            "amount": amount,
        })

    close_fn()
    return records
