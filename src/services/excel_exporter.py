"""Export settlement data to Excel format with formulas for calculated cells."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_settlement_to_excel(
    driver_name: str,
    start_date: str,
    end_date: str,
    results: list[dict],
    totals: dict,
    fuel_deducted: bool = False,
) -> BytesIO:
    """Generate Excel file from settlement data.

    Calculated cells use Excel formulas instead of static values.
    DATA cells (from CSV/config) have blue font.
    FORMULA cells have default black font.

    Args:
        fuel_deducted: If True, anticipado formula subtracts fuel (col N).

    Returns BytesIO buffer containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidacion"

    # Title rows
    ws["A1"] = f"Liquidacion: {driver_name}"
    ws["A2"] = f"Periodo: {start_date} - {end_date}"
    ws["A1"].font = Font(bold=True, size=14)

    # Column headers in row 4
    headers = [
        "Fecha", "Prima", "Inc", "FreenowT3", "Uber T3",
        "Rec. Neta", "%", "TPV", "App FN", "App Uber",
        "Cash", "IVA", "Parte Prop.",
        "Gasolina", "Otros", "Anticipado", "Liquidacion",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Styles
    blue_font = Font(color="0000FF")
    num_fmt = '#,##0.00'

    # DATA columns (1-indexed): B=2, C=3, D=4, E=5, G=7, H=8, I=9, J=10, N=14, O=15
    data_cols = {2, 3, 4, 5, 7, 8, 9, 10, 14, 15}
    # FORMULA columns: F=6, K=11, L=12, M=13, P=16, Q=17
    formula_cols = {6, 11, 12, 13, 16, 17}

    # Data rows
    row = 5
    for r in results:
        # A: Fecha (no number format)
        ws.cell(row=row, column=1, value=r["date"].strftime("%d/%m/%Y"))

        # DATA cells — write values with blue font
        ws.cell(row=row, column=2, value=float(r.get("prima_amount", 0)))
        ws.cell(row=row, column=3, value=float(r.get("incidents_amount", 0)))
        ws.cell(row=row, column=4, value=float(r.get("freenow_fixed", 0)))
        ws.cell(row=row, column=5, value=float(r.get("uber_t3_fixed", 0)))
        ws.cell(row=row, column=7, value=float(r.get("driver_pct", 0)))
        ws.cell(row=row, column=8, value=float(r.get("tpv_visa_total", 0)))
        ws.cell(row=row, column=9, value=float(r.get("freenow_app", 0)))
        ws.cell(row=row, column=10, value=float(r.get("uber_total_payment", 0)))
        ws.cell(row=row, column=14, value=float(r.get("fuel_total", 0)))
        ws.cell(row=row, column=15, value=float(r.get("other_expenses_total", 0)))

        # FORMULA cells
        # F: Rec. Neta = B + D + E - C
        ws.cell(row=row, column=6, value=f"=B{row}+D{row}+E{row}-C{row}")
        # K: Cash = F - H - I - J
        ws.cell(row=row, column=11, value=f"=F{row}-H{row}-I{row}-J{row}")
        # L: IVA = F - F/1.1
        ws.cell(row=row, column=12, value=f"=F{row}-F{row}/1.1")
        # M: Parte Prop. = (F - L) * G / 100
        ws.cell(row=row, column=13, value=f"=(F{row}-L{row})*G{row}/100")
        # P: Anticipado = K - O (or K - O - N if fuel_deducted)
        if fuel_deducted:
            ws.cell(row=row, column=16, value=f"=K{row}-O{row}-N{row}")
        else:
            ws.cell(row=row, column=16, value=f"=K{row}-O{row}")
        # Q: Liquidacion = M - P
        ws.cell(row=row, column=17, value=f"=M{row}-P{row}")

        # Apply number format and font color
        for col in range(2, 18):
            cell = ws.cell(row=row, column=col)
            cell.number_format = num_fmt
            if col in data_cols:
                cell.font = blue_font

        row += 1

    # Totals row with SUM formulas
    last_data_row = row - 1
    total_row = row
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True)

    ws.cell(row=total_row, column=1, value="TOTAL").font = total_font

    # All numeric columns get SUM formulas (skip A=date, G=percentage)
    sum_cols = [2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
    for col in sum_cols:
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=total_row, column=col,
            value=f"=SUM({col_letter}5:{col_letter}{last_data_row})",
        )
        cell.fill = total_fill
        cell.font = total_font
        cell.number_format = num_fmt

    # Column widths
    col_widths = [12, 10, 10, 12, 10, 12, 6, 10, 10, 12, 10, 10, 10, 10, 10, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
