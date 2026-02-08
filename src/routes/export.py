# src/routes/export.py
import csv
from datetime import date
from io import BytesIO, StringIO
from sqlalchemy import func, cast, Date
from sqlalchemy.orm import Session
from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from src.routes.auth import require_admin
from src.database import get_session
from src.models.trip import Trip
from src.models.driver import Driver
from src.models.vehicle import Vehicle
from src.template_config import templates

router = APIRouter()


@router.get("/export", response_class=HTMLResponse)
async def export_page(
    request: Request,
    user: dict = Depends(require_admin),
    session: Session = Depends(get_session),
):
    drivers = session.query(Driver).filter_by(is_active=True).all()
    return templates.TemplateResponse(request, "export.html", {
        "user": user,
        "drivers": drivers,
    })


def _query_trips(session: Session, date_from: date, date_to: date, driver_id: str | None):
    q = (
        session.query(Trip)
        .filter(cast(Trip.started_at, Date) >= date_from)
        .filter(cast(Trip.started_at, Date) <= date_to)
        .order_by(Trip.started_at)
    )
    if driver_id:
        q = q.filter(Trip.driver_id == driver_id)
    return q.all()


def _query_daily_summary(session: Session, date_from: date, date_to: date, driver_id: str | None):
    q = (
        session.query(
            cast(Trip.started_at, Date).label("date"),
            Trip.driver_id,
            func.count(Trip.id).label("trip_count"),
            func.coalesce(func.sum(Trip.gross_amount), 0).label("gross"),
            func.coalesce(func.sum(Trip.commission), 0).label("commission"),
            func.coalesce(func.sum(Trip.payout_amount), 0).label("net"),
            func.coalesce(func.sum(Trip.distance_km), 0).label("km"),
        )
        .filter(cast(Trip.started_at, Date) >= date_from)
        .filter(cast(Trip.started_at, Date) <= date_to)
        .group_by(cast(Trip.started_at, Date), Trip.driver_id)
        .order_by(cast(Trip.started_at, Date))
    )
    if driver_id:
        q = q.filter(Trip.driver_id == driver_id)
    return q.all()


def _trips_to_csv(trips, drivers_map, vehicles_map) -> str:
    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Fecha", "Hora", "Plataforma", "Conductor", "Vehiculo",
        "Bruto", "Comision", "IVA", "Propinas", "Peajes", "Neto",
        "Distancia_km", "Duracion_min", "Metodo_pago", "Tarifa",
    ])
    for t in trips:
        writer.writerow([
            t.started_at.strftime("%d/%m/%Y") if t.started_at else "",
            t.started_at.strftime("%H:%M") if t.started_at else "",
            t.source,
            drivers_map.get(t.driver_id, "?"),
            vehicles_map.get(t.vehicle_id, "?"),
            f"{float(t.gross_amount):.2f}",
            f"{float(t.commission):.2f}",
            f"{float(t.taxes_vat):.2f}",
            f"{float(t.tips):.2f}",
            f"{float(t.tolls):.2f}",
            f"{float(t.payout_amount or 0):.2f}",
            f"{float(t.distance_km or 0):.2f}",
            f"{float(t.duration_minutes or 0):.1f}",
            t.payment_method or "",
            t.tariff_code or "",
        ])
    return output.getvalue()


def _summary_to_csv(rows, drivers_map) -> str:
    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Fecha", "Conductor", "Viajes", "Bruto", "Comision", "Neto", "Km"])
    for r in rows:
        writer.writerow([
            r.date,
            drivers_map.get(r.driver_id, "?"),
            r.trip_count,
            f"{float(r.gross):.2f}",
            f"{float(r.commission):.2f}",
            f"{float(r.net):.2f}",
            f"{float(r.km):.1f}",
        ])
    return output.getvalue()


def _trips_to_excel(trips, drivers_map, vehicles_map) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    headers = [
        "Fecha", "Hora", "Plataforma", "Conductor", "Vehiculo",
        "Bruto", "Comision", "IVA", "Propinas", "Peajes", "Neto",
        "Km", "Min", "Pago", "Tarifa",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for i, t in enumerate(trips, 2):
        ws.cell(row=i, column=1, value=t.started_at.strftime("%d/%m/%Y") if t.started_at else "")
        ws.cell(row=i, column=2, value=t.started_at.strftime("%H:%M") if t.started_at else "")
        ws.cell(row=i, column=3, value=t.source)
        ws.cell(row=i, column=4, value=drivers_map.get(t.driver_id, "?"))
        ws.cell(row=i, column=5, value=vehicles_map.get(t.vehicle_id, "?"))
        ws.cell(row=i, column=6, value=float(t.gross_amount))
        ws.cell(row=i, column=7, value=float(t.commission))
        ws.cell(row=i, column=8, value=float(t.taxes_vat))
        ws.cell(row=i, column=9, value=float(t.tips))
        ws.cell(row=i, column=10, value=float(t.tolls))
        ws.cell(row=i, column=11, value=float(t.payout_amount or 0))
        ws.cell(row=i, column=12, value=float(t.distance_km or 0))
        ws.cell(row=i, column=13, value=float(t.duration_minutes or 0))
        ws.cell(row=i, column=14, value=t.payment_method or "")
        ws.cell(row=i, column=15, value=t.tariff_code or "")
        for col in range(6, 14):
            ws.cell(row=i, column=col).number_format = '#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _summary_to_excel(rows, drivers_map) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    headers = ["Fecha", "Conductor", "Viajes", "Bruto", "Comision", "Neto", "Km"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=str(r.date))
        ws.cell(row=i, column=2, value=drivers_map.get(r.driver_id, "?"))
        ws.cell(row=i, column=3, value=r.trip_count)
        ws.cell(row=i, column=4, value=float(r.gross))
        ws.cell(row=i, column=5, value=float(r.commission))
        ws.cell(row=i, column=6, value=float(r.net))
        ws.cell(row=i, column=7, value=float(r.km))
        for col in range(4, 8):
            ws.cell(row=i, column=col).number_format = '#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/export/download")
async def download_export(
    request: Request,
    user: dict = Depends(require_admin),
    session: Session = Depends(get_session),
    date_from: str = Form(...),
    date_to: str = Form(...),
    format: str = Form("csv"),
    driver_id: str = Form(""),
    report_type: str = Form("trips"),
):
    df = date.fromisoformat(date_from)
    dt = date.fromisoformat(date_to)
    did = driver_id if driver_id else None

    drivers_map = {d.id: d.name for d in session.query(Driver).all()}
    vehicles_map = {v.id: v.plate for v in session.query(Vehicle).all()}

    filename_base = f"taxi_{report_type}_{date_from}_{date_to}"

    if report_type == "summary":
        rows = _query_daily_summary(session, df, dt, did)
        if format == "excel":
            buf = _summary_to_excel(rows, drivers_map)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"},
            )
        else:
            content = _summary_to_csv(rows, drivers_map)
            return StreamingResponse(
                iter([content]),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"},
            )
    else:
        # trips or tax (tax = trips with all fields)
        trips = _query_trips(session, df, dt, did)
        if format == "excel":
            buf = _trips_to_excel(trips, drivers_map, vehicles_map)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"},
            )
        else:
            content = _trips_to_csv(trips, drivers_map, vehicles_map)
            return StreamingResponse(
                iter([content]),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"},
            )
